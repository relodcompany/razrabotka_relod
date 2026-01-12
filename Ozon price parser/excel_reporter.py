# excel_reporter.py
import os
from datetime import datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

# Глобальные объекты книги/листа — одна книга на один запуск анализа
_WB: Optional[Workbook] = None
_WS: Optional[Worksheet] = None
_INITIALIZED: bool = False

# Заголовки и порядок колонок строго по ТЗ
_HEADERS = [
    "Название товара",
    "ISBN",
    "Наша цена",
    "Конкурент 1",
    "Цена конкурента 1",
    "Конкурент 2",
    "Цена конкурента 2",
    "Конкурент 3",
    "Цена конкурента 3",
]


def _ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _init_workbook_once():
    global _WB, _WS, _INITIALIZED
    if _INITIALIZED and _WB is not None and _WS is not None:
        return
    _ensure_output_dir()
    _WB = Workbook()
    _WS = _WB.active
    _WS.title = "Отчёт"
    # Пишем шапку
    _WS.append(_HEADERS)
    # Стили ширины колонок — чтобы было читабельно
    widths = {
        1: 60,  # Название товара
        2: 20,  # ISBN
        3: 14,  # Наша цена
        4: 36,  # Конкурент 1
        5: 18,  # Цена конкурента 1
        6: 36,  # Конкурент 2
        7: 18,  # Цена конкурента 2
        8: 36,  # Конкурент 3
        9: 18,  # Цена конкурента 3
    }
    for col_idx, w in widths.items():
        _WS.column_dimensions[get_column_letter(col_idx)].width = w
    _INITIALIZED = True


def _fmt_price(val) -> str:
    """Нормализуем цену к строке (без валюты), если это число или строка с числом."""
    if val is None:
        return ""
    try:
        # чаще всего мы передаём price как float/int или как строку вида '866 ₽'
        if isinstance(val, (int, float)):
            return str(int(val)) if float(val).is_integer() else f"{float(val):.2f}"
        s = str(val).strip()
        # выкидываем пробелы и знак валюты, если вдруг пролез
        s = s.replace("₽", "").replace("\u2009", "").strip()
        # если это число — ок
        float(s.replace(",", "."))
        # показываем как есть (без валюты)
        return s
    except Exception:
        return str(val)


def write_competitor_report_row(
    *,
    isbn: str,
    title: str,
    our_price: Optional[float],
    competitors: List[Dict],
) -> None:
    """
    Добавляет строку в Excel-отчёт.
    competitors — список словарей, каждый может содержать:
      - 'shop' (или 'seller') — название магазина
      - 'price' и/или 'gray_price' — числовая/строковая цена
      - 'url' — ссылка (не выводим в отчёт, но можно добавить при необходимости)

    Берём до 3 конкурентов с минимальными ценами (подразумевается, что передан уже отсортированный top_n),
    а если конкурентов меньше — заполняем «нет конкурента».
    """
    _init_workbook_once()
    assert _WS is not None

    # Берём до 3 конкурентов
    # Порядок — как пришли (find_top_competitor_offers_by_isbn уже отдаёт top_n по возрастанию цены)
    comps = competitors[:3] if competitors else []

    # Подготовка полей
    name = title or ""
    isbn_str = isbn or ""
    our_price_str = _fmt_price(our_price) if our_price is not None else ""

    row = [name, isbn_str, our_price_str]

    # Добавляем до 3 конкурентов
    for idx in range(3):
        if idx < len(comps):
            c = comps[idx]
            shop = c.get("shop") or c.get("seller") or ""
            # для цены предпочитаем 'gray_price', если есть (как просили, "серая цена")
            price_val = c.get("gray_price", None)
            if price_val in (None, "", 0):
                price_val = c.get("price", "")
            row.extend([shop, _fmt_price(price_val)])
        else:
            row.extend(["нет конкурента", ""])

    _WS.append(row)


def save_workbook() -> str:
    """
    Сохраняет книгу в OUTPUT_DIR и возвращает путь к файлу.
    Имя файла: competitor_prices_YYYYmmdd_HHMMSS.xlsx
    """
    _init_workbook_once()
    assert _WB is not None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"competitor_prices_{ts}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    _WB.save(path)
    return path
